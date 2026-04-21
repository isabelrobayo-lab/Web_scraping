"""Unit tests for export generation in CSV, Excel, and JSON formats.

Tests the export service (generate_export) and verifies that each format
produces valid output files with correct content.
"""

import json
import os
import tempfile
from decimal import Decimal
from unittest.mock import patch

import pytest

# Patch settings before importing the service
_test_export_dir = tempfile.mkdtemp()


@pytest.fixture(autouse=True)
def _patch_export_dir():
    """Patch EXPORT_DIR to use a temp directory for all tests."""
    with patch("app.exports.service.settings") as mock_settings:
        mock_settings.EXPORT_DIR = _test_export_dir
        yield


def _sample_records(n: int = 5) -> list[dict]:
    """Generate sample property records for testing."""
    records = []
    for i in range(n):
        records.append({
            "id_interno": i + 1,
            "codigo_inmueble": f"INM-{i + 1:04d}",
            "sitio_origen": "test-site.com",
            "tipo_inmueble": "Apartamento",
            "operacion": "Venta",
            "municipio": "Bogotá",
            "precio_local": Decimal("250000000.50"),
            "metros_utiles": Decimal("85.5"),
            "habitaciones": 3,
            "banos": 2,
            "estado_activo": True,
            "titulo_anuncio": f"Apartamento {i + 1}",
        })
    return records


class TestGenerateExportCSV:
    """Tests for CSV export generation."""

    def test_csv_creates_file(self):
        """CSV export creates a .csv file on disk."""
        from app.exports.service import generate_export

        records = _sample_records(3)
        file_path, count = generate_export(records, "csv", "test-csv-001")

        assert os.path.exists(file_path)
        assert file_path.endswith(".csv")
        assert count == 3

    def test_csv_content_has_headers_and_rows(self):
        """CSV file contains headers and the correct number of data rows."""
        from app.exports.service import generate_export

        records = _sample_records(2)
        file_path, _ = generate_export(records, "csv", "test-csv-002")

        with open(file_path, "r", encoding="utf-8") as f:
            lines = f.readlines()

        # Header + 2 data rows
        assert len(lines) == 3
        header = lines[0].strip()
        assert "codigo_inmueble" in header
        assert "sitio_origen" in header

    def test_csv_empty_records(self):
        """CSV export with empty records creates a file with 0 records."""
        from app.exports.service import generate_export

        file_path, count = generate_export([], "csv", "test-csv-empty")

        assert os.path.exists(file_path)
        assert count == 0


class TestGenerateExportExcel:
    """Tests for Excel export generation."""

    def test_excel_creates_file(self):
        """Excel export creates an .xlsx file on disk."""
        from app.exports.service import generate_export

        records = _sample_records(3)
        file_path, count = generate_export(records, "excel", "test-excel-001")

        assert os.path.exists(file_path)
        assert file_path.endswith(".xlsx")
        assert count == 3

    def test_excel_content_readable(self):
        """Excel file can be read back with openpyxl and has correct rows."""
        from app.exports.service import generate_export

        records = _sample_records(4)
        file_path, _ = generate_export(records, "excel", "test-excel-002")

        import openpyxl

        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        # Header row + 4 data rows
        assert ws.max_row == 5
        # Check header contains expected column
        headers = [cell.value for cell in ws[1]]
        assert "codigo_inmueble" in headers

    def test_excel_empty_records(self):
        """Excel export with empty records creates a valid file."""
        from app.exports.service import generate_export

        file_path, count = generate_export([], "excel", "test-excel-empty")

        assert os.path.exists(file_path)
        assert count == 0


class TestGenerateExportJSON:
    """Tests for JSON export generation."""

    def test_json_creates_file(self):
        """JSON export creates a .json file on disk."""
        from app.exports.service import generate_export

        records = _sample_records(3)
        file_path, count = generate_export(records, "json", "test-json-001")

        assert os.path.exists(file_path)
        assert file_path.endswith(".json")
        assert count == 3

    def test_json_content_is_valid(self):
        """JSON file contains valid JSON with correct record count."""
        from app.exports.service import generate_export

        records = _sample_records(2)
        file_path, _ = generate_export(records, "json", "test-json-002")

        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)

        assert isinstance(data, list)
        assert len(data) == 2
        assert data[0]["codigo_inmueble"] == "INM-0001"
        assert data[1]["codigo_inmueble"] == "INM-0002"

    def test_json_empty_records(self):
        """JSON export with empty records creates a valid empty array."""
        from app.exports.service import generate_export

        file_path, count = generate_export([], "json", "test-json-empty")

        assert os.path.exists(file_path)
        assert count == 0

        with open(file_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        assert data == []


class TestGenerateExportValidation:
    """Tests for export format validation."""

    def test_invalid_format_raises_error(self):
        """Unsupported format raises ValueError."""
        from app.exports.service import generate_export

        with pytest.raises(ValueError, match="no soportado"):
            generate_export([], "xml", "test-invalid")


class TestExportHelpers:
    """Tests for export helper functions."""

    def test_get_content_type_csv(self):
        from app.exports.service import get_content_type

        assert get_content_type("csv") == "text/csv"

    def test_get_content_type_excel(self):
        from app.exports.service import get_content_type

        assert "spreadsheetml" in get_content_type("excel")

    def test_get_content_type_json(self):
        from app.exports.service import get_content_type

        assert get_content_type("json") == "application/json"

    def test_get_file_extension(self):
        from app.exports.service import get_file_extension

        assert get_file_extension("csv") == ".csv"
        assert get_file_extension("excel") == ".xlsx"
        assert get_file_extension("json") == ".json"
